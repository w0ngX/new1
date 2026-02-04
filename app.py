# -*- coding: utf-8 -*-
import os
import json
import io
import pandas as pd
import httpx
import urllib.parse
import calendar
from sqlalchemy.orm import joinedload
from datetime import date
from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, jsonify, Response, \
    stream_with_context, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import or_, func,and_
# app.py 顶部引入
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openai import OpenAI

# --- 1. 数据库驱动配置 ---
import pymysql

# 让 SQLAlchemy 以为自己在用标准驱动 (解决 MySQL 连接问题)
pymysql.install_as_MySQLdb()

# --- 2. Flask 应用初始化 ---
app = Flask(__name__)
app.secret_key = 'your_secret_key_here_root'  # 请修改为安全的密钥
# 1. 定义你的原始密码
#raw_password = "Root@123456"
raw_password = "root"
# 2. 对密码进行转义处理
safe_password = urllib.parse.quote_plus(raw_password)
# 3. 拼接到连接字符串中
# 注意：密码位置换成变量 safe_password
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:{}@localhost:3306/data_system?charset=utf8mb4'.format(safe_password)
# 数据库连接配置 (请确认密码是否正确)
#app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:Root%40123456@localhost:5636/data_system?charset=utf8mb4'
#app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:root@localhost:3306/data_system?charset=utf8mb4'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# 文件上传配置
# 文件上传配置
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf', 'doc', 'docx', 'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# 初始化插件
db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = '请先登录系统！'
login_manager.login_message_category = 'warning'

# 兼容性处理：创建干净的 http 客户端
try:
    # 尝试新版 httpx 的语法
    custom_http_client = httpx.Client(proxies=None)
except TypeError:
    # 如果报错，说明是旧版 httpx，使用 proxy 参数
    custom_http_client = httpx.Client(proxy=None)
# --- 3. AI 配置 (通义千问) ---
client = OpenAI(
    api_key="sk-ab4860410e584a708300e40ae4985159",  # 您的 Key
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
    http_client=custom_http_client  # ✨ 关键：强行使用这个“干净”的客户端
)


# --- 4. 数据库模型定义 (关键修正部分) ---

# 用户模型
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)

    def set_password(self, password):
        #self.password_hash = generate_password_hash(password)
        # 强制指定加密方法为 pbkdf2，这是所有系统都支持的
        self.password_hash = generate_password_hash(password, method='pbkdf2:sha256')

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# 交付记录模型 (唯一且正确)
class DeliveryLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    # 外键关联
    contract_id = db.Column(db.Integer, db.ForeignKey('contract.id'), nullable=False)
    delivery_date = db.Column(db.String(20), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    related_unit = db.Column(db.String(200))
    created_at = db.Column(db.DateTime, default=db.func.now())


# 合同模型
class Contract(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    contract_code = db.Column(db.String(50), nullable=True)
    name = db.Column(db.String(200), nullable=False)
    year = db.Column(db.String(10))
    party_a = db.Column(db.String(100), nullable=True)
    party_b = db.Column(db.String(100), nullable=True)
    amount = db.Column(db.Float, default=0.0)
    balance = db.Column(db.Float, default=0.0)
    sign_date = db.Column(db.String(20), nullable=True)
    end_date = db.Column(db.String(20), nullable=True)
    contract_type = db.Column(db.String(50))  # 新增：合同類型
    scan_file = db.Column(db.String(500))
    manual_status = db.Column(db.String(20), nullable=True)
    # 新增/确认以下两个金额字段
    #total_amount = db.Column(db.Float, default=0.0)  # 合同总金额
    #paid_amount = db.Column(db.Float, default=0.0)  # 已付金额 (您提到的"余额")

    # 【关键修正】关联关系指向 DeliveryLog
    # cascade="all, delete-orphan" 保证删除合同时自动删除交付记录，不报错
    deliveries = db.relationship('DeliveryLog', backref='contract', lazy=True, cascade="all, delete-orphan")

    @property
    def is_settled(self):
        # 已结清：余额 <= 0 或 余额 < 0.01（考虑浮点数误差）
        return self.balance is not None and self.balance <= 0.01
        # 如果 余额 >= 总金额，则视为已结清
        # (加个 > 0 判断防止空合同被误判)
        #return self.balance >= self.amount and self.amount > 0

    @property
    def paid_amount(self):
        """已支付金额"""
        # 已支付金额 = 总金额 - 余额
        if self.amount is not None and self.balance is not None:
            return max(0, self.amount - self.balance)
        return 0

    # 在 Contract 模型中修改 update_balance 方法
    def update_balance(self):
        """根据交付记录更新余额"""
        # 计算所有交付金额之和
        total_delivered = db.session.query(func.sum(DeliveryLog.amount)).filter_by(contract_id=self.id).scalar() or 0
        # 余额 = 总金额 - 已交付金额
        self.balance = max(0, self.amount - total_delivered)
        return self.balance


# === 批量导入支付记录功能 ===

@app.route('/import_payments_excel', methods=['POST'])
@login_required
def import_payments_excel():
    file = request.files.get('file')
    if not file:
        flash("请选择要上传的Excel文件", "warning")
        return redirect(url_for('contracts'))

    try:
        # 读取Excel文件，不自动推断数据类型
        df = pd.read_excel(file, dtype=str).replace({pd.NA: None, float('nan'): None})

        # 检查必要的列是否存在
        required_cols = ['合同编号', '支付金额', '支付时间']
        if not all(col in df.columns for col in required_cols):
            missing_cols = [col for col in required_cols if col not in df.columns]
            # flash(f"Excel文件缺少必要的列：{', '.join(missing_cols)}", "danger")
            # 将第151行改为：
            message = "Excel文件缺少必要的列：" + ", ".join(missing_cols)
            flash(message, "danger")
            return redirect(url_for('contracts'))

        success_count = 0
        fail_count = 0
        fail_messages = []

        # 统计信息
        summary = {
            'updated_contracts': set(),
            'total_amount': 0.0
        }

        for index, row in df.iterrows():
            try:
                contract_code_raw = str(row.get('合同编号') or '').strip()
                payment_amount_raw = row.get('支付金额')
                payment_date_raw = row.get('支付时间')

                # 验证合同编号
                if not contract_code_raw:
                    # fail_messages.append(f"第{index + 2}行：合同编号为空")
                    fail_messages.append("第{}行：合同编号为空".format(index + 2))
                    fail_count += 1
                    continue

                # 处理合同编号：去除多余空格和特殊字符
                contract_code = contract_code_raw.strip()

                # 验证支付金额
                try:
                    payment_amount = float(str(payment_amount_raw).replace(',', ''))
                    if payment_amount <= 0:
                        #fail_messages.append(f"第{index + 2}行：支付金额必须大于0")
                        fail_messages.append("第{}行：支付金额'{}'无效，支付金额必须大于0".format(index + 2, payment_amount_raw))
                        fail_count += 1
                        continue
                except (ValueError, TypeError):
                    # fail_messages.append(f"第{index + 2}行：支付金额'{payment_amount_raw}'无效")
                    fail_messages.append("第{}行：支付金额'{}'无效".format(index + 2,payment_amount_raw))
                    fail_count += 1
                    continue

                # 处理支付日期
                payment_date_str = None
                if payment_date_raw:
                    payment_date_str = force_clean_date(payment_date_raw)

                if not payment_date_str:
                    # 如果没有有效日期，使用固定日期
                    payment_date_str = '2023-01-01'

                # 根据合同编号查找合同（多种匹配方式）
                contract = None

                # 方式1：完全匹配合同编号
                contract = Contract.query.filter_by(contract_code=contract_code).first()

                # 方式2：如果合同编号是数字，尝试作为ID查找
                if not contract and contract_code.isdigit():
                    try:
                        contract_id = int(contract_code)
                        contract = Contract.query.get(contract_id)
                    except:
                        pass

                # 方式3：在合同名称中查找
                if not contract:
                    contracts = Contract.query.filter(
                        db.or_(
                            Contract.name.contains(contract_code),
                            Contract.contract_code.contains(contract_code)
                        )
                    ).all()

                    if len(contracts) == 1:
                        contract = contracts[0]
                    elif len(contracts) > 1:
                        # 如果有多个匹配，记录警告但使用第一个
                        # fail_messages.append(f"第{index + 2}行：合同编号'{contract_code}'匹配到多个合同，使用第一个")
                        fail_messages.append("第{}行：合同编号'{}'匹配到多个合同，使用第一个".format(index + 2, contract_code))
                        contract = contracts[0]

                if not contract:
                    #fail_messages.append(f"第{index + 2}行：未找到合同编号为'{contract_code}'的合同")
                    fail_messages.append("第{}行：未找到合同编号为'{}'的合同".format(index + 2, contract_code))
                    fail_count += 1
                    continue

                # 检查是否已存在相同的支付记录（防止重复导入）
                existing_payment = DeliveryLog.query.filter_by(
                    contract_id=contract.id,
                    delivery_date=payment_date_str,
                    amount=payment_amount
                ).first()

                if existing_payment:
                   #fail_messages.append(
                   #     f"第{index + 2}行：支付记录已存在（合同：{contract.name}，日期：{payment_date_str}，金额：{payment_amount}）")
                    fail_messages.append(
                        "第{}行：支付记录已存在（合同：{}，日期：{}，金额：{}）".format(index + 2, contract.name,
                                                                                payment_date_str, payment_amount))
                    fail_count += 1
                    continue

                # 创建支付记录
                delivery_log = DeliveryLog(
                    contract_id=contract.id,
                    delivery_date=payment_date_str,
                    amount=payment_amount,
                    related_unit=row.get('支付单位', '').strip() or row.get('相关单位', '').strip()
                )

                db.session.add(delivery_log)

                # 更新统计信息
                summary['updated_contracts'].add(contract.id)
                summary['total_amount'] += payment_amount

                success_count += 1

            except Exception as e:
                #fail_messages.append(f"第{index + 2}行：处理失败 - {str(e)}")
                fail_messages.append("第{}行：处理失败 - {}".format(index + 2, str(e)))
                fail_count += 1
                continue

        # 提交事务
        db.session.commit()

        # 重新计算所有受影响合同的余额
        for contract_id in summary['updated_contracts']:
            contract = Contract.query.get(contract_id)
            if contract:
                contract.update_balance()

        db.session.commit()

        # 构建反馈消息
        if success_count > 0:
            #success_msg = f"✅ 支付记录导入完成！成功导入 {success_count} 条记录"
            success_msg = "✅ 支付记录导入完成！成功导入 {} 条记录".format(success_count)
            if summary['updated_contracts']:
                #success_msg += f"，更新了 {len(summary['updated_contracts'])} 个合同的余额"
                success_msg += "，更新了 {} 个合同的余额".format(len(summary['updated_contracts']))
            if summary['total_amount'] > 0:
                #success_msg += f"，总支付金额 ¥{summary['total_amount']:,.2f}"
                success_msg += "，总支付金额 ¥{:,.2f}".format(summary['total_amount'])

            flash(success_msg, "success")

        if fail_count > 0:
            #fail_summary = f"⚠️  有 {fail_count} 条记录导入失败"
            fail_summary = "⚠️  有 {} 条记录导入失败".format(fail_count)
            #if fail_messages:
             #   # 显示前5条详细错误
              #  fail_summary += f"<br><small>"
               # for i, msg in enumerate(fail_messages[:5]):
                #    fail_summary += f"{i + 1}. {msg}<br>"
                #if fail_count > 5:
                 #   #fail_summary += f"... 还有 {fail_count - 5} 条错误未显示"
                  #  fail_summary += "... 还有 {} 条错误未显示".format(fail_count - 5)
                  #  fail_summary += "</small>"

            flash(fail_summary, "warning")

        return redirect(url_for('contracts'))

    except Exception as e:
        db.session.rollback()
        #flash(f"❌ 导入出错：{str(e)}", "danger")
        flash("❌ 导入出错：{}".format(str(e)), "danger")
        return redirect(url_for('contracts'))


# 主题模型
class Topic(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    level1_id = db.Column(db.String(50), nullable=False)
    level2_id = db.Column(db.String(50), nullable=False)
    theme_id = db.Column(db.String(50), nullable=False)
    name = db.Column(db.String(200), nullable=False)
    method = db.Column(db.String(100), nullable=True)
    frequency = db.Column(db.String(50), nullable=True)
    owner = db.Column(db.String(100), nullable=True)


# 任务模型
class Task(db.Model):
    __tablename__ = 'task'
    id = db.Column(db.Integer, primary_key=True)
    contract_name = db.Column(db.String(200))
    service_content = db.Column(db.Text)
    theme_name = db.Column(db.String(200))
    contract_id = db.Column(db.Integer, db.ForeignKey('contract.id'))
    # 👇 关键是这一句！它让你可以通过 task.contract 访问整个合同对象
    contract = db.relationship('Contract', backref='tasks')


# --- 5. 辅助函数 ---

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def force_clean_date(value):
    """
    极强健壮性的日期清洗函数，支持多种格式
    """
    if value is None or str(value).strip().lower() in ['none', 'nan', '', 'nat']:
        return None

    # 1. 如果本身就是 datetime 对象（pandas 自动识别的情况）
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d')

    # 2. 处理"2025-09"这样的年月格式
    val_str = str(value).strip()

    # 匹配格式：YYYY-MM 或 YYYY/MM 或 YYYY.MM
    import re
    if re.match(r'^\d{4}[-/.]\d{1,2}$', val_str):
        try:
            # 如果是年月格式，自动添加01作为日
            if '-' in val_str:
                year, month = val_str.split('-')
            elif '/' in val_str:
                year, month = val_str.split('/')
            else:
                year, month = val_str.split('.')

            # 确保月份是两位数
            month = month.zfill(2)

            # 如果是单个数字月份（如2025-9），补零
            if len(month) == 1:
                month = '0' + month
            #return f"{year}-{month}-01"
            return "{}-{}-01".format(year,month)
        except:
            pass

    # 3. 如果是字符串，尝试多种格式解析
    formats = [
        '%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y',
        '%Y.%m.%d', '%Y%m%d', '%Y-%m-%d %H:%M:%S'
    ]

    for fmt in formats:
        try:
            return datetime.strptime(val_str, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue

    # 4. 如果是 Excel 的数字日期格式（例如 45123）
    try:
        if val_str.replace('.', '').isdigit():
            # Excel 基准日期是 1899-12-30
            excel_date = datetime(1899, 12, 30) + timedelta(days=float(val_str))
            return excel_date.strftime('%Y-%m-%d')
    except:
        pass

    return None


def get_all_owners():
    default_owners = {'兼职', '李忠科', '唐小语', '张楷雯', '赵黛莉','(未指定负责人)'}
    db_owners = db.session.query(Topic.owner).filter(Topic.owner != None, Topic.owner != '').distinct().all()
    current_owners = {r[0] for r in db_owners}
    return sorted(list(default_owners.union(current_owners)))


# 【核心修复】这里补全了 is_active 函数
@app.context_processor
def inject_globals():
    def is_active(endpoint):
        if request.endpoint and request.endpoint == endpoint: return 'active'
        # 让子页面也能点亮父级菜单
        if endpoint == 'contracts' and request.endpoint in ['contracts', 'edit_contract']: return 'active'
        if endpoint == 'topics' and request.endpoint in ['topics', 'edit_topic']: return 'active'
        if endpoint == 'tasks' and request.endpoint in ['tasks', 'edit_task']: return 'active'
        return ''

    #return dict(is_active=is_active, today=date.today().strftime('%Y-%m-%d'))
    return dict(is_active=is_active, today='2023-01-01')


# --- 6. 路由定义 ---

# === 认证模块 ===
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            #flash(f'欢迎回来，{user.username}！', 'success')
            flash('欢迎回来，{}！'.format(user.username), 'success')
            return redirect(url_for('index'))
        else:
            flash('账号或密码错误', 'danger')
    return render_template('login.html', title="管理员登录")


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('您已安全退出', 'info')
    return redirect(url_for('login'))


# === 首页总览 ===
# --- app.py 修改 index 函数 ---
@app.route('/')
@login_required
def index():
    contract_count = Contract.query.count()
    topic_count = Topic.query.count()
    task_count = Task.query.count()

    # 强制固定日期界限
    today_str = '2023-01-01'
    #today = today_str
    # 统计逻辑：考虑手动状态 + 固定日期
    active_contracts = Contract.query.filter(
        db.or_(
            Contract.manual_status == 'active',
            db.and_(
                Contract.manual_status == None,
                Contract.end_date != None,
                Contract.end_date != '',
                Contract.end_date >= today_str
            )
        )
    ).count()
    ended_contracts = contract_count - active_contracts

    active_percent = round((active_contracts / contract_count) * 100, 1) if contract_count > 0 else 0
    ended_percent = round((ended_contracts / contract_count) * 100, 1) if contract_count > 0 else 0
    #重点合同筛选
    #important_contracts_query = Contract.query.filter(
    #    Contract.party_a.like('%税%'),  # 甲方含“税”
    #    or_(
    #        Contract.manual_status == 'active',
    #        #request.args.get('status') == 'active'
    #    )
    #)

    #important_contracts = important_contracts_query.all()
    #important_count = important_contracts_query.count()
    # 修改后的逻辑：必须满足条件1，并且满足条件2或条件3之一
    important_contracts = Contract.query.filter(
        # 条件1：甲方必须包含"稅"
        #Contract.party_a.like('%税%'),
        Contract.contract_type.like('%税务%'),
        # 条件2或条件3（只要满足一个即可）
        db.or_(
            # 条件2：手动标记为active
            Contract.manual_status == 'active',
            # 条件3：未手动标记且未过期
            db.and_(
                Contract.manual_status == None,
                Contract.end_date != None,
                Contract.end_date != '',
                Contract.end_date >= today_str
            )
        )
    ).all()
    important_count = len(important_contracts)
    jsb_contracts = Contract.query.filter(
        # 条件1：甲方必须包含"稅"
        Contract.party_b.like('%极数宝%'),
        Contract.contract_type.like('%税务%'),
        # 条件2或条件3（只要满足一个即可）
        db.or_(
            # 条件2：手动标记为active
            Contract.manual_status == 'active',
            # 条件3：未手动标记且未过期
            db.and_(
                Contract.manual_status == None,
                Contract.end_date != None,
                Contract.end_date != '',
                Contract.end_date >= today_str
            )
        )
    ).all()
    jsb_count = len(jsb_contracts)
    mf_contracts = Contract.query.filter(
        # 条件1：甲方必须包含"稅"
        Contract.party_b.like('%蚂蜂%'),
        Contract.contract_type.like('%税务%'),
        # 条件2或条件3（只要满足一个即可）
        db.or_(
            # 条件2：手动标记为active
            Contract.manual_status == 'active',
            # 条件3：未手动标记且未过期
            db.and_(
                Contract.manual_status == None,
                Contract.end_date != None,
                Contract.end_date != '',
                Contract.end_date >= today_str
            )
        )
    ).all()
    mf_count = len(mf_contracts)
    try:
        # 1. 基礎統計：合同總數與總金額
        total_contracts = Contract.query.count()
        total_amount = db.session.query(func.sum(Contract.amount)).scalar() or 0

        # 2. 待付款統計：以「餘額 (balance) > 0」作為未結清標準
        # 如果你的資料庫是用 is_settled 欄位，請將 filter 條件改回 Contract.is_settled == 0
        unsettled_query = Contract.query.filter(Contract.balance > 0)
        unsettled_count = unsettled_query.count()
        unsettled_amount = db.session.query(func.sum(Contract.balance)).filter(Contract.balance > 0).scalar() or 0

        # 3. 最近合同
        recent_contracts = Contract.query.order_by(Contract.id.desc()).limit(5).all()

        return render_template('index.html',
                               contract_count=contract_count, topic_count=topic_count, task_count=task_count,
                               active_contracts=active_contracts, active_percent=active_percent,
                               ended_contracts=ended_contracts, ended_percent=ended_percent,
                               total_contracts=total_contracts,
                               total_amount=total_amount,
                               unsettled_count=unsettled_count,
                               unsettled_amount=unsettled_amount,
                               recent_contracts=recent_contracts,
                               important_contracts=important_contracts,
                               important_count=important_count,
                               jsb_contracts=jsb_contracts,
                               jsb_count=jsb_count,
                               mf_contracts=mf_contracts,
                               mf_count=mf_count,
                               )
    except Exception as e:
            print(f"首頁加載出錯: {str(e)}")
            # 出錯時給予默認值，避免 500 錯誤
            return render_template('index.html',
                                   total_contracts=0, total_amount=0,
                                   unsettled_count=0, unsettled_amount=0,
                                   recent_contracts=[])
# === AI 接口 ===
@app.route('/api/ai_query', methods=['POST'])
@login_required
def ai_query():
    data = request.get_json()
    user_input = data.get('query', '').strip() if data else ''
    if not user_input: return jsonify({"error": "问题不能为空"}), 400

    # --- 核心：多维度知识构建 ---
    knowledge_chunks = []

    # 1. 提取合同与财务大纲
    contracts = Contract.query.all()
    knowledge_chunks.append("=== 合同与财务概况 ===")
    for c in contracts:
        status = "履约中" if (c.end_date and c.end_date >= '2023-01-01') else "已结束"
        if c.manual_status: status = "履约中" if c.manual_status == 'active' else "已结束"
        #knowledge_chunks.append(
        #    f"合同:{c.name}, 金额:{c.amount}, 余额:{c.balance}, 状态:{status}, 编号:{c.contract_code}")
        knowledge_chunks.append("合同:{}, 金额:{}, 余额:{}, 状态:{}, 编号:{}".format(
            c.name, c.amount, c.balance, status, c.contract_code))

    # 2. 提取主题与负责人映射
    topics = Topic.query.all()
    knowledge_chunks.append("\n=== 数据主题与负责架构 ===")
    for t in topics:
        #knowledge_chunks.append(f"主题:{t.name}, 负责人:{t.owner}, 采集频率:{t.frequency}, 方式:{t.method}")
        knowledge_chunks.append("主题:{}, 负责人:{}, 采集频率:{}, 方式:{}".format(
            t.name, t.owner, t.frequency, t.method))

    # 3. 提取服务内容明细 (关联映射)
    tasks = Task.query.all()
    knowledge_chunks.append("\n=== 服务内容明细记录 ===")
    for k in tasks:
        #knowledge_chunks.append(
        #    f"在合同[{k.contract_name}]下, 提供了服务:[{k.service_content}], 涉及主题:[{k.theme_name}]")
        knowledge_chunks.append("在合同[{}]下, 提供了服务:[{}], 涉及主题:[{}]".format(
            k.contract_name, k.service_content, k.theme_name))

    # 合并为完整的上下文
    full_knowledge = "\n".join(knowledge_chunks)

    # --- 调用 AI ---
    def generate():
        try:
            response = client.chat.completions.create(
                model="qwen-plus",
                messages=[
                    {
                        "role": "system",
                        #"content": f"你是一个大数据管理系统的专家。你的知识库如下：\n{full_knowledge}\n请根据以上信息，综合分析并回答用户问题。如果涉及统计，请给出具体数字。回答要专业且简洁。"
                        "content": "你是一个大数据管理系统的专家。你的知识库如下：\n{}\n请根据以上信息，综合分析并回答用户问题。如果涉及统计，请给出具体数字。回答要专业且简洁。".format(full_knowledge)

            },
                    {"role": "user", "content": user_input}
                ],
                stream=True
            )
            for chunk in response:
                if chunk.choices and chunk.choices[0].delta.content:
                    #yield f"data: {json.dumps({'text': chunk.choices[0].delta.content}, ensure_ascii=False)}\n\n"
                    yield "data: {}\n\n".format(
                        json.dumps({'text': chunk.choices[0].delta.content}, ensure_ascii=False)
                    )

            yield "data: [DONE]\n\n"
        except Exception as e:
            #yield f"data: {json.dumps({'text': f'系统解析出错: {str(e)}'})}\n\n"
            yield "data: {}\n\n".format(
                json.dumps({'text': '系统解析出错: {}'.format(str(e))})
            )

    return Response(stream_with_context(generate()), mimetype='text/event-stream')


# === 合同管理 ===
# --- app.py 修改 contracts 函数 ---
@app.route('/contracts')
@login_required
def contracts():
    page = request.args.get('page', 1, type=int)
    # ... (获取参数代码) ...
    # --- 必须确保以下参数获取代码存在 ---
    s_code = request.args.get('code', '')  # 新增：合同编号
    s_name = request.args.get('name', '')
    s_contract_type = request.args.get('contract_type', '')
    s_party_a = request.args.get('party_a', '')
    s_party_b = request.args.get('party_b', '')
    s_type = request.args.get('type', '')
    s_status = request.args.get('status', '').strip()  # 执行状态 (active/ended)
    s_min_paid = request.args.get('min_paid', type=float)
    s_max_paid = request.args.get('max_paid', type=float)
    s_pay_status = request.args.get('pay_status', '').strip()  # 🚨 新增：结算状态 (settled/unsettled)
    # ----------------------------------
    # 强制固定日期界限
    today_str = '2023-01-01'

    # 顶部条形图统计逻辑
    total_count = Contract.query.count()
    active_count = Contract.query.filter(
        db.or_(
            Contract.manual_status == 'active',
            db.and_(
                Contract.manual_status == None,
                Contract.end_date != None,
                Contract.end_date != '',
                Contract.end_date >= today_str
            )
        )
    ).count()
    ended_count = total_count - active_count

    active_percent = round((active_count / total_count) * 100, 1) if total_count else 0
    ended_percent = round((ended_count / total_count) * 100, 1) if total_count else 0

    # 列表筛选逻辑
    query = Contract.query
    # ... (名称、单位筛选代码) ...

    if s_status == 'active':
        query = query.filter(
            db.or_(
                Contract.manual_status == 'active',
                db.and_(Contract.manual_status == None, Contract.end_date >= today_str)
            )
        )
    elif s_status == 'ended':
        query = query.filter(
            db.or_(
                Contract.manual_status == 'ended',
                db.and_(Contract.manual_status == None,
                        db.or_(Contract.end_date < today_str, Contract.end_date == None))
            )
        )

    pagination = query.order_by(Contract.id.desc()).paginate(page=page, per_page=50)
    # 1. 获取搜索参数
    search_query = request.args.get('q', '').strip()
    status_filter = request.args.get('status', '').strip()  # 获取状态筛选

    # 2. 基础查询
    query = Contract.query

    # 1. 关键词搜索 (原有逻辑)
    if search_query:
        query = query.filter(Contract.contract_name.contains(search_query))
        # 1. 关键词筛选
        # 名称与单位筛选
    if s_code:
        # 使用模糊搜索：合同编号或合同名称中包含搜索词
        query = query.filter(
            db.or_(
                Contract.contract_code.contains(s_code),
                Contract.name.contains(s_code)  # 同时搜索合同名称，因为有些合同可能没有编号
            )
        )
    if s_name:
        query = query.filter(Contract.name.contains(s_name))
    if s_contract_type:  # 新增：根据合同类型过滤
        query = query.filter(Contract.contract_type == s_contract_type)

    if s_party_a:
        query = query.filter(Contract.party_a.contains(s_party_a))
    if s_party_b:
        query = query.filter( Contract.party_b.contains(s_party_b))

    # 执行状态筛选 (解决您报错的 if/elif 块)
    if s_status == 'active':
        query = query.filter(
            db.or_(
                Contract.manual_status == 'active',
                db.and_(Contract.manual_status == None, Contract.end_date >= today_str)
            )
        )
    elif s_status == 'ended':
        query = query.filter(
            db.or_(
                Contract.manual_status == 'ended',
                db.and_(Contract.manual_status == None,
                        db.or_(Contract.end_date < today_str, Contract.end_date == None))
            )
        )
    # 在 contracts() 函数中修改结算状态筛选
    if s_pay_status == 'settled':
        # 已结清：余额 <= 0.01
        #query = query.filter(Contract.balance > Contract.amount)
        query = query.filter(Contract.balance <= 0.01)
    elif s_pay_status == 'unsettled':
        # 未结清：余额 > 0.01
        #query = query.filter(Contract.balance <= Contract.amount)
        query = query.filter(Contract.balance > 0.01)
    # 获取数据库中所有已存在的合同类型（去重并过滤空值）
    all_types = [t[0] for t in db.session.query(Contract.contract_type).distinct().all() if t[0]]
    #type_options = [t[0] for t in all_types if t[0]]
    type_query = db.session.query(Contract.contract_type).distinct().all()
    type_options = [t[0] for t in type_query if t[0] and t[0].strip()]
    page = request.args.get('page', 1, type=int)
    pagination = query.order_by(Contract.id.desc()).paginate(page=page, per_page=50)
    return render_template('contracts.html',
                           title="合同管理",
                           pagination=pagination,
                           active_count=active_count,
                           ended_count=ended_count,
                           active_percent=active_percent,
                           ended_percent=ended_percent,
                           today=today_str,
                           status_filter=s_pay_status,  # 结算状态回显 (对应前端 request.args.get('pay_status')),
                           search_status=s_status,
                           search_name=s_name,
                           search_type=s_type,
                           search_party_a=s_party_a,
                           search_party_b=s_party_b,
                           search_code=s_code,
                           all_types=all_types,
                           type_options=type_options
                           )


@app.route('/contract/edit/<int:id>', methods=['GET', 'POST'])
@app.route('/contract/add', methods=['GET', 'POST'], defaults={'id': None})
@login_required
def edit_contract(id):
    contract = Contract.query.get_or_404(id) if id else None
    # 获取数据库中已存在的去重后的合同类型
    all_types = db.session.query(Contract.contract_type).distinct().all()
    all_types = [t[0] for t in all_types if t[0]]  # 转换为列表并过滤空值

    if request.method == 'POST':
        if not contract:
            contract = Contract()
            # 新增合同：余额初始等于总金额
            #contract.balance = float(request.form.get('amount', 0))
            #contract.balance = 0.0  # <--- 初始化为 0
            db.session.add(contract)

        contract.contract_code = request.form.get('contract_code')
        contract.name = request.form.get('name')
        contract.year = request.form.get('year'),
        contract.party_a = request.form.get('party_a')
        contract.party_b = request.form.get('party_b')
        contract.sign_date = request.form.get('sign_date')
        contract.end_date = request.form.get('end_date')
        #contract.amount = float(request.form.get('amount', 0))
        contract.contract_type = request.form.get('contract_type')  # 獲取類型
        amount_input = float(request.form.get('amount', 0))
        contract.amount = amount_input
        # 余额默认为总金额（新建时）
        #if not id: contract.balance = contract.amount
        if not id:
            contract.balance = amount_input
        else:
            # 如果是编辑现有合同，保持现有余额逻辑
            # 但需要确保金额变化时余额合理
            if contract.amount != amount_input:
                # 如果总金额改变了，需要重新计算余额
                contract.update_balance()
        # --- 核心：扫描件逻辑处理 ---
        # 1. 获取上传的文件
        file = request.files.get('scan_file')
        # 2. 获取填写的链接
        link = request.form.get('scan_link', '').strip()

        if file and file.filename != '' and allowed_file(file.filename):
            # 优先处理文件上传
            #filename = secure_filename(f"{contract.contract_code or 'TEMP'}_{file.filename}")
            filename = secure_filename("{}_{}".format(contract.contract_code or 'TEMP', file.filename))
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            contract.scan_file = filename
        elif link:
            # 如果没传文件但填了链接，则保存链接
            contract.scan_file = link
        # ---------------------------

        db.session.commit()
        flash('合同保存成功！', 'success')
        return redirect(url_for('contracts'))

    return render_template('contract_form.html',
                           contract=contract,
                           all_types=all_types)

#合同类型更新
@app.route('/api/update_contract_type', methods=['POST'])
@login_required
def update_contract_type_api():
    try:
        data = request.json
        c_id = data.get('id')
        new_type = data.get('type')

        contract = db.session.get(Contract, c_id)
        if contract:
            contract.contract_type = new_type
            db.session.commit()
            return jsonify({'status': 'success'})
        return jsonify({'status': 'error', 'message': '找不到合同'}), 404
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/contract/delete/<int:id>')
@login_required
def delete_contract(id):
    contract = Contract.query.get_or_404(id)
    db.session.delete(contract)
    db.session.commit()
    flash('合同已删除', 'danger')
    return redirect(url_for('contracts'))


@app.route('/contract/<int:id>/delete_file', methods=['POST'])
@login_required
def delete_contract_file(id):
    contract = Contract.query.get_or_404(id)

    if contract.scan_file and not contract.scan_file.startswith('http'):
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], contract.scan_file)

        try:
            # 1. 从磁盘删除文件
            if os.path.exists(file_path):
                os.remove(file_path)

            # 2. 清空数据库记录
            contract.scan_file = None
            db.session.commit()

            return jsonify({"status": "success", "message": "文件已成功删除"})
        except Exception as e:
            db.session.rollback()
            #return jsonify({"status": "error", "message": f"删除失败: {str(e)}"}), 500
            return jsonify({"status": "error", "message": "删除失败: {}".format(str(e))}), 500

    return jsonify({"status": "error", "message": "未找到可删除的本地文件"}), 400


# --- 交付记录管理 (修正版) ---

@app.route('/contract/<int:id>/delivery_history')
@login_required
def get_delivery_history(id):
    contract = Contract.query.get_or_404(id)
    logs = DeliveryLog.query.filter_by(contract_id=id).order_by(DeliveryLog.delivery_date.desc()).all()
    total_delivered = sum(log.amount for log in logs)

    return jsonify({
        "contract_name": contract.name,
        "total_amount": contract.amount,
        "total_delivered": total_delivered,
        "current_balance": contract.balance,  # 添加当前余额
        "is_settled": contract.is_settled,  # 添加是否已结清状态
        "history": [
            {
                "id": log.id,
                "date": log.delivery_date,
                "amount": log.amount,
                "unit": log.related_unit
            } for log in logs
        ]
    })


@app.route('/contract/<int:id>/add_delivery', methods=['POST'])
@login_required
def add_delivery(id):
    contract = Contract.query.get_or_404(id)
    delivery_date = request.form.get('delivery_date')
    amount = float(request.form.get('amount', 0))
    related_unit = request.form.get('related_unit')  # 这里对应前端的“说明/单位”

    if amount > 0:
        # 1. 新增记录
        new_log = DeliveryLog(
            contract_id=id,
            delivery_date=delivery_date,
            amount=amount,
            related_unit=related_unit
        )
        db.session.add(new_log)
        db.session.commit()  # 先提交，保证记录入库

        # 2. 【核心】重新计算余额 (Sum 所有记录)
        # 这样能保证余额永远等于所有支付记录之和，绝对准确
        #total_paid = db.session.query(func.sum(DeliveryLog.amount)).filter_by(contract_id=id).scalar() or 0
        #contract.balance = total_paid
        # 2. 【核心修改】更新余额：余额 = 余额 - 交付金额
        # 但余额不能小于0
        #if contract.balance is None:
        #   contract.balance = contract.amount

        #contract.balance = max(0, contract.balance - amount)
        new_balance = contract.update_balance()
        db.session.commit()  # 再次提交更新余额

        return {"status": "success", "new_balance": contract.balance, "is_settled": contract.is_settled}

    return {"status": "error", "message": "金额必须大于0"}, 400


@app.route('/delivery/delete/<int:id>', methods=['POST'])
@login_required
def delete_delivery(id):
    delivery = DeliveryLog.query.get_or_404(id)
    contract_id = delivery.contract_id
    contract = Contract.query.get(contract_id)

    if not contract:
        return jsonify({'status': 'error', 'message': '合同不存在'}), 404

    try:
        # 记录删除前的余额用于反馈
        old_balance = contract.balance

        # 删除交付记录
        db.session.delete(delivery)

        # 重新计算余额
        new_balance = contract.update_balance()
        db.session.commit()

        return jsonify({
            'status': 'success',
            #'message': f'删除成功，余额从 ¥{old_balance:,.2f} 更新为 ¥{new_balance:,.2f}',
            'message': '删除成功，余额从 ¥{:,.2f} 更新为 ¥{:,.2f}'.format(old_balance,new_balance),
            'new_balance': new_balance
        })

    except Exception as e:
        db.session.rollback()
        #return jsonify({'status': 'error', 'message': f'系统错误: {str(e)}'}), 500
        return jsonify({'status': 'error', 'message': '系统错误: {}'.format(str(e))}), 500


@app.route('/contract/<int:id>/update_status', methods=['POST'])
@login_required
def update_contract_status(id):
    contract = Contract.query.get_or_404(id)
    new_status = request.json.get('status')
    contract.manual_status = new_status if new_status in ['active', 'ended'] else None
    db.session.commit()
    return jsonify({"status": "success"})


# === 数据校准工具 ===
@app.route('/admin/fix_balances')
@login_required
def fix_balances():
    contracts = Contract.query.all()
    count = 0

    for contract in contracts:
        # 使用新的 update_balance 方法重新计算余额
        old_balance = contract.balance
        new_balance = contract.update_balance()

        if abs(old_balance - new_balance) > 0.01:  # 考虑浮点数误差
            count += 1
            #print(f"合同 {contract.name}: 余额从 {old_balance} 修正为 {new_balance}")
            print("合同 {}: 余额从 {} 修正为 {}".format(contract.name,old_balance,new_balance))

    db.session.commit()
    return "校准完成！修复了 {} 个合同的余额。<a href='/contracts'>返回</a>".format(count)


# --- 新增：导入合同余额Excel文件 ---
@app.route('/import_balance_excel', methods=['POST'])
@login_required
def import_balance_excel():
    """终极版导入：自动处理缺少xlrd的情况，支持双重ID匹配"""
    file = request.files.get('file')
    if not file:
        flash("请选择要上传的文件", "warning")
        return redirect(url_for('contracts'))

    try:
        filename = file.filename.lower()
        print("正在处理文件: {}".format(file.filename))

        # --- 1. 智能文件读取 (带容错处理) ---
        try:
            if filename.endswith('.csv'):
                try:
                    df = pd.read_csv(file, dtype=str)
                except UnicodeDecodeError:
                    file.seek(0)
                    df = pd.read_csv(file, dtype=str, encoding='gbk')
            else:
                # Excel 读取
                #df = pd.read_excel(file, dtype=str)
                # 找到读取 Excel 的地方，改为这样：
                df = pd.read_excel(file)

                # 1. 强制清洗所有列名的空格
                df.columns = [str(c).strip() for c in df.columns]

                # 2. 将所有数据先转为字符串，防止数字类型在云端数据库冲突
                # 这一步能大幅提高导入成功率
                df = df.fillna('')  # 把空值填为空字符串，防止 NaN 报错

                #print(f"DEBUG: 正在尝试导入 {len(df)} 行数据...")
                print("DEBUG: 成功读取 Excel，共 {} 行数据".format(len(df)))

        except ImportError as e:
            if 'xlrd' in str(e):
                flash(
                    "❌ 系统缺少读取 .xls 文件的库。请运行 'pip install xlrd' 安装，或将文件另存为 .xlsx / .csv 格式上传。",
                    "danger")
                return redirect(url_for('contracts'))
            raise e
        except Exception as e:
            #flash(f"❌ 文件读取失败，请检查文件格式。错误: {str(e)}", "danger")
            flash("❌ 文件读取失败，请检查文件格式。错误: {}".format(str(e)), "danger")
            return redirect(url_for('contracts'))

        # 2. 数据清洗
        df.columns = [str(c).strip() for c in df.columns]
        df = df.replace({pd.NA: None, float('nan'): None})

        # 3. 智能列名映射
        # 自动寻找可能的列名
        col_contract_code = next((c for c in df.columns if c in ['合同编号', '编号', 'Code']), None)
        col_uuid = next((c for c in df.columns if c in ['HTMXID', 'UUID', 'ID']), None)
        col_amount = next((c for c in df.columns if c in ['支付金额', '金额', 'Amount']), None)
        col_date = next((c for c in df.columns if c in ['支付时间', '日期', 'Time', 'Date']), None)
        col_remark = next((c for c in df.columns if c in ['HTMXBZ', '备注', 'Remark']), None)

        if not col_amount:
            #flash(f"失败：未找到[支付金额]列。检测到的列名：{list(df.columns)}", "danger")
            flash("失败：未找到[支付金额]列。检测到的列名：{}".format(list(df.columns)), "danger")
            return redirect(url_for('contracts'))

        # 只要有其中一种编号即可
        if not (col_contract_code or col_uuid):
            #flash(f"失败：未找到[合同编号]或[HTMXID]列。", "danger")
            flash("失败：未找到[合同编号]列。", "danger")
            return redirect(url_for('contracts'))

        success_count = 0
        fail_count = 0
        fail_msgs = []
        updated_contract_ids = set()

        for index, row in df.iterrows():
            try:
                # --- A. 获取与清洗数据 ---
                money_raw = row.get(col_amount)
                if pd.isna(money_raw): continue

                # 金额清洗
                try:
                    amount = float(str(money_raw).replace(',', '').replace('¥', '').strip())
                except:
                    continue

                # 日期清洗
                date_raw = row.get(col_date)
                try:
                    if date_raw:
                        # 自动解析日期
                        dt = pd.to_datetime(date_raw)
                        pay_date = dt.strftime('%Y-%m-%d')
                    else:
                        pay_date = '2023-01-01'
                except:
                    pay_date = '2023-01-01'

                # --- B. 匹配合同 (核心优化) ---
                contract = None

                # 策略1：优先尝试用 HTMXID (长ID) 匹配
                if col_uuid:
                    uuid_val = str(row.get(col_uuid) or '').strip()
                    if uuid_val:
                        contract = Contract.query.filter_by(contract_code=uuid_val).first()

                # 策略2：尝试用 合同编号 (短编号) 匹配
                if not contract and col_contract_code:
                    code_val = str(row.get(col_contract_code) or '').strip()
                    if code_val:
                        contract = Contract.query.filter_by(contract_code=code_val).first()
                        # 策略3：尝试模糊匹配名称
                        if not contract:
                            contract = Contract.query.filter(Contract.name.contains(code_val)).first()

                if not contract:
                    display_code = row.get(col_contract_code) or row.get(col_uuid) or '未知'
                    #fail_msgs.append(f"行{index + 2}: 编号[{display_code}]未找到对应合同")
                    fail_msgs.append("行{}: 编号[{}]未找到对应合同".format(index + 2, display_code))
                    fail_count += 1
                    continue

                # --- C. 查重与入库 ---
                exists = DeliveryLog.query.filter_by(
                    contract_id=contract.id,
                    delivery_date=pay_date,
                    amount=amount
                ).first()

                if exists:
                    continue

                remark = str(row.get(col_remark) or '').strip()
                log = DeliveryLog(
                    contract_id=contract.id,
                    delivery_date=pay_date,
                    amount=amount,
                    related_unit=remark
                )
                db.session.add(log)
                updated_contract_ids.add(contract.id)
                success_count += 1

            except Exception as e:
              #  fail_msgs.append(f"行{index + 2}: 异常 {str(e)}")
                fail_msgs.append("行{}: 异常 {}".format(index + 2, str(e)))
                fail_count += 1

        db.session.commit()

        # --- D. 刷新余额 ---
        for cid in updated_contract_ids:
            c = Contract.query.get(cid)
            if c:
                c.update_balance()
        db.session.commit()

        # --- E. 反馈结果 ---
        if success_count > 0:
            #flash(f"✅ 成功导入 {success_count} 条，更新 {len(updated_contract_ids)} 个合同余额。", "success")
            flash("✅ 成功导入 {} 条，更新 {} 个合同余额。".format(success_count, len(updated_contract_ids)), "success")

        if fail_count > 0:
            err_str = "<br>".join(fail_msgs[:3])
            #flash(f"⚠️ {fail_count} 条失败。<br>前3条原因：<br>{err_str}", "warning")
            flash("⚠️ {} 条失败。<br>前3条原因：<br>{}".format(fail_count, err_str), "warning")
        elif success_count == 0:
            flash("⚠️ 未导入任何数据，请检查文件内容。", "warning")

        return redirect(url_for('contracts'))

    except Exception as e:
        db.session.rollback()
        #print(f"Server Error: {e}")
        print("Server Error: {}".format(e))
        #flash(f"❌ 系统错误: {str(e)}", "danger")
        flash("❌ 系统错误: {}".format(str(e)), "danger")
        return redirect(url_for('contracts'))


# --- 新增：手动匹配合同功能（用于调试） ---
@app.route('/debug_match_contract/<contract_code>')
@login_required
def debug_match_contract(contract_code):
    """调试合同匹配功能"""
    contracts = Contract.query.all()
    matches = []

    for contract in contracts:
        # 检查各种匹配方式
        if contract.contract_code == contract_code:
            matches.append("精确匹配合同编号: {} (编号: {})".format(contract.name,contract.contract_code))

        if contract_code in contract.name:
            matches.append("合同名称包含编号: {} (编号: {})".format(contract.name,contract.contract_code))

        if contract.contract_code and contract_code in contract.contract_code:
            matches.append("合同编号包含: {} (编号: {})".format(contract.name,contract.contract_code))

    #return "<br>".join(matches) if matches else f"未找到匹配合同编号: {contract_code}"
    return "<br>".join(matches) if matches else "未找到匹配合同编号: {}".format(contract_code)

# === 导入功能 ===
@app.route('/import_excel', methods=['POST'])
@login_required
def import_excel():
    file = request.files.get('file')
    if not file:
        flash("请选择要上传的 Excel 文件", "warning")
        return redirect(url_for('contracts'))

    try:
        # 读取 Excel 并处理空值
        df = pd.read_excel(file).replace({pd.NA: None, float('nan'): None})

        new_count = 0
        update_count = 0

        for _, row in df.iterrows():
            name = str(row.get('合同名称') or '').strip()
            code = str(row.get('合同编号') or '').strip()

            if not name:
                continue

            # 1. 尝试查找数据库中是否已存在该编号的合同
            existing_contract = None
            if code:
                existing_contract = Contract.query.filter_by(contract_code=code).first()

            if existing_contract:
                # 2. 如果存在：执行更新操作
                existing_contract.name = name
                existing_contract.contract_type = str(row.get('合同类型') or '').strip()
                existing_contract.party_a = str(row.get('甲方单位') or '').strip()
                existing_contract.party_b = str(row.get('乙方单位') or '').strip()
                existing_contract.year = str(row.get('所属年度') or '').strip()
                existing_contract.amount = float(row.get('合同金额') or 0)
                # 注意：余额同步更新，如果需要保留手动修改的余额，可以删掉下面这行
                existing_contract.balance = float(row.get('合同金额') or 0)
                existing_contract.sign_date = force_clean_date(row.get('签约日期'))
                existing_contract.end_date = force_clean_date(row.get('履约日期'))
                existing_contract.update_balance()
                # 安全更新：只有当 Excel 提供了链接时才更新扫描件，防止覆盖掉已有的本地文件
                scan_link = str(row.get('扫描件链接') or '').strip()
                if scan_link:
                    existing_contract.scan_file = scan_link

                update_count += 1
            else:
                amount_val = float(row.get('合同金额') or 0)
                # 3. 如果不存在：执行新增操作
                new_c = Contract(
                    contract_code=code,

                    name=name,
                    contract_type=str(row.get('合同类型') or '').strip(),
                    party_a=str(row.get('甲方单位') or '').strip(),
                    party_b=str(row.get('乙方单位') or '').strip(),
                    year=str(row.get('所属年度') or '').strip(),
                    amount=amount_val,
                    balance=amount_val,
                    sign_date=force_clean_date(row.get('签约日期')),
                    end_date=force_clean_date(row.get('履约日期')),
                    scan_file=str(row.get('扫描件链接') or '').strip() or None
                )
                db.session.add(new_c)
                new_count += 1

        db.session.commit()

        # 4. 反馈详细的结果
        #msg = f"导入完成！成功新增 {new_count} 条数据，更新 {update_count} 条现有数据。"
        msg = "导入完成！成功新增 {} 条数据，更新 {} 条现有数据。".format(new_count, update_count), "success"
        flash(msg, "success")
        return redirect(url_for('contracts'))

    except Exception as e:
        db.session.rollback()
      #  flash(f"导入出错：{str(e)}", "danger")
        flash("导入出错：{}".format(str(e)), "danger")
        return redirect(url_for('contracts'))
#导出合同Excel

@app.route('/admin/init_balances')
@login_required
def init_balances():
    """初始化所有合同的余额（用于从旧系统迁移）"""
    contracts = Contract.query.all()
    updated = 0

    for contract in contracts:
        if contract.balance is None:
            # 如果余额为空，设置为总金额
            contract.balance = contract.amount or 0
            updated += 1
        else:
            # 如果已有余额，但需要确保逻辑正确
            # 计算所有交付金额
            total_delivered = db.session.query(func.sum(DeliveryLog.amount)).filter_by(
                contract_id=contract.id).scalar() or 0
            # 预期余额 = 总金额 - 已交付金额
            expected_balance = max(0, (contract.amount or 0) - total_delivered)

            if abs(contract.balance - expected_balance) > 0.01:
                contract.balance = expected_balance
                updated += 1
                #print(f"合同 {contract.name}: 余额从 {contract.balance} 修正为 {expected_balance}")
                print("合同 {}: 余额从 {} 修正为 {}".format(contract.name, contract.balance, expected_balance))

    db.session.commit()
    return "余额初始化完成！更新了 {} 个合同。<a href='/contracts'>返回</a>".format(updated)


@app.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


# === 主题管理 ===
@app.route('/topics')
@login_required
def topics():
    page = request.args.get('page', 1, type=int)
    s_name = request.args.get('name', '').strip()
    s_parent = request.args.get('parent', '').strip()
    s_level2 = request.args.get('level2', '').strip()
    s_level2_select = request.args.get('level2_select', '').strip()
    s_owner_select = request.args.get('owner_select', '').strip()  # 新增：下拉筛选参数
    s_owner = request.args.get('owner_select', '').strip()

    # 在后端查询所有不重复的组合
    raw_relations = db.session.query(Topic.level1_id, Topic.level2_id, Topic.name).distinct().all()
    # 转换为前端易处理的列表
    # 格式: [{'p': '一级1', 'l2': '二级1', 'n': '主题1'}, ...]
    topic_relations = [{"p": r[0] or "", "l2": r[1] or "", "n": r[2] or ""} for r in raw_relations]

    # 1. 提取所有去重后的列表数据
    all_parents = [t[0] for t in db.session.query(Topic.level1_id).distinct().all() if t[0]]
    all_level2_list = [t[0] for t in db.session.query(Topic.level2_id).distinct().all() if t[0]]
    raw_owners = db.session.query(Topic.owner).distinct().all()
    all_owners = [t[0] for t in raw_owners if t[0] and t[0].strip()]
    #all_owners = [t[0] for t in db.session.query(Topic.owner).distinct().all() if t[0]]  # 负责人列表
    # 2. 注入“虚拟负责人”到列表首位，方便前端循环
    all_owners.insert(0, "(未指定负责人)")
    s_owner_select = request.args.get('owner', '').strip()
    # 2. 构建查询
    query = Topic.query
    if s_name:
        query = query.filter(Topic.name.contains(s_name))
    if s_parent:
        query = query.filter(Topic.level1_id == s_parent)
    if s_level2_select:
        query = query.filter(Topic.level2_id == s_level2_select)
    if s_owner_select:
        #query = query.filter(Topic.owner == s_owner_select)  # 精确匹配负责人
        if s_owner_select == "(未指定负责人)":  # 定义一个特殊的标识符
               # 匹配 负责人字段为 NULL 或 空字符串 的记录
            query = query.filter(or_(Topic.owner == None, Topic.owner == '', Topic.owner == '(未指定负责人)'))
        else:
            query = query.filter(Topic.owner == s_owner_select)
    if s_level2:
        query = query.filter(Topic.level2_id == s_level2)
    if s_owner:
        query = query.filter(Topic.owner == s_owner)

    pagination = query.order_by(Topic.id.desc()).paginate(page=page, per_page=50)

    return render_template('topics.html',
                           pagination=pagination,
                           all_parents=all_parents,
                           all_level2_list=all_level2_list,
                           all_owners=all_owners,
                           s_parent=s_parent,
                           s_level2_select=s_level2_select,
                           s_owner_select=s_owner_select,
                           topic_relations=topic_relations,
                           s_name=s_name)  # 传回当前选中的负责人


# 在 app.py 中找到 edit_topic 函数并完全替换为以下内容：

@app.route('/topic/edit/<int:id>', methods=['GET', 'POST'])
@app.route('/topic/add', methods=['GET', 'POST'], defaults={'id': None})
@login_required
def edit_topic(id):
    topic = Topic.query.get_or_404(id) if id else None

    if request.method == 'POST':
        # ... (保存逻辑保持不变) ...
        if not topic:
            topic = Topic()
            db.session.add(topic)
        topic.level1_id = request.form.get('level1_id')
        topic.level2_id = request.form.get('level2_id')
        topic.theme_id = request.form.get('theme_id')
        topic.name = request.form.get('name')
        topic.method = ",".join(request.form.getlist('methods'))
        topic.frequency = request.form.get('frequency')

        # ✨ 核心修改：如果为空，自动填入 (未指定负责人)
        owner_input = request.form.get('owner', '').strip()
        topic.owner = owner_input if owner_input else "(未指定负责人)"
        #topic.owner = request.form.get('owner', '').strip() or "(未指定负责人)"
        #topic.owner = request.form.get('owner')
        db.session.commit()
        flash('主题保存成功', 'success')
        return redirect(url_for('topics'))

    # ========================================================
    # 🔍 暴力调试 + 强制数据填充区
    # ========================================================

    # 1. 查询所有【一级主题】(过滤掉空值)
    raw_level1 = db.session.query(Topic.level1_id).distinct().all()
    all_parents = [r[0] for r in raw_level1 if r[0] and r[0].strip()]

    # 🚨 强制兜底：如果没查到，给个默认值，证明下拉栏是好的
    if not all_parents:
        all_parents = ["示例一级A", "示例一级B (数据库无数据)"]

    # 2. 查询所有【二级主题】
    raw_level2 = db.session.query(Topic.level2_id).distinct().all()
    all_level2_list = [r[0] for r in raw_level2 if r[0] and r[0].strip()]

    if not all_level2_list:
        all_level2_list = ["示例二级X", "示例二级Y (数据库无数据)"]

    # 3. 查询所有【负责人】
    # 结合 get_all_owners 函数（确保该函数在 app.py 里定义了）
    try:
        all_owners = get_all_owners()
    except:
        # 万一函数没定义，直接查库
        raw_owners = db.session.query(Topic.owner).distinct().all()
        all_owners = [r[0] for r in raw_owners if r[0] and r[0].strip()]

    if not all_owners:
        all_owners = ["张三", "李四 (数据库无数据)"]

    # ========================================================

    return render_template('topic_form.html',
                           topic=topic,
                           # 确保变量名和模板里 {% for p in all_parents %} 一致
                           all_parents=all_parents,
                           all_level2_list=all_level2_list,
                           all_owners=all_owners)


@app.route('/topic/delete/<int:id>')
@login_required
def delete_topic(id):
    db.session.delete(Topic.query.get_or_404(id))
    db.session.commit()
    return redirect(url_for('topics'))


@app.route('/import_topic_excel', methods=['POST'])
@login_required
def import_topic_excel():
    file = request.files.get('file')
    if not file: return "无文件", 400
    try:
        df = pd.read_excel(file).replace({pd.NA: None, float('nan'): None})
        for _, row in df.iterrows():
            t_name = str(row.get('主题名称') or '').strip()
            if not t_name: continue
            if Topic.query.filter_by(name=t_name).first(): continue  # 简易查重

            raw_method = str(row.get('采集方式') or '')
            methods = []
            if "程序" in raw_method: methods.append("程序采集")
            if "人工" in raw_method: methods.append("人工采集")

            # --- 获取负责人数据 ---
            # --- 获取负责人数据 (优化版) ---
            raw_owner_val = row.get('负责人')

            # 严谨的空值判断：处理 None, NaN, 空字符串, 纯空格, 以及字符串 "None"
            if (raw_owner_val is None or
                    pd.isna(raw_owner_val) or
                    str(raw_owner_val).strip() == '' or
                    str(raw_owner_val).lower() in ['nan', 'none']):

                final_owner = "(未指定负责人)"
            else:
                final_owner = str(raw_owner_val).strip()
            # --- 执行入库 ---
            # 无论是有名字还是"(未指定负责人)"，都平等地存入数据库
            topic = Topic.query.filter_by(id=row.get('ID')).first()
            if topic:
                topic.owner = final_owner

            new_t = Topic(
                level1_id=str(row.get('一级主题') or ''),
                level2_id=str(row.get('二级主题') or ''),
                theme_id=str(row.get('主题ID') or ''),
                name=t_name,
                method=",".join(methods),
                frequency=str(row.get('采集频率') or ''),
                owner=final_owner # 👈 使用处理后的 final_owner
            )
            db.session.add(new_t)
        db.session.commit()
        flash('导入完成', 'success')
        return redirect(url_for('topics'))
    except Exception as e:
        #return f"错误: {e}", 500
        return "错误: {}".format(e), 500


# app.py (放在 download_topic_template 附近)
#Excel列宽识别
def get_visual_length(text):
    """
    精准计算文本的视觉长度：
    - 中文/全角符号：算 2 个单位
    - 英文/数字：算 1 个单位
    """
    if not text:
        return 0
    text = str(text)
    length = 0
    for char in text:
        # 判断是否为汉字或全角符号 (根据Unicode编码范围)
        if '\u4e00' <= char <= '\u9fff' or '\uff00' <= char <= '\uffef':
            length += 2
        else:
            length += 1
    return length

#Excel 美化
def beautify_excel(writer, df, sheet_name='Sheet1'):
    """
    Excel 美化终极版：
    1. 智能列宽：遍历数据，取最大值
    2. 视觉优化：表头深蓝背景 + 白字 + 边框
    """
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    # --- 样式定义 ---
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    content_font = Font(name='微软雅黑')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')  # 内容建议左对齐或居中，看喜好

    # --- 1. 设置表头 (第一行) ---
    for col_num, value in enumerate(df.columns.values):
        cell = worksheet.cell(row=1, column=col_num + 1)
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    worksheet.row_dimensions[1].height = 25  # 表头稍微高一点

    # --- 2. 智能列宽计算 ---
    for i, col_name in enumerate(df.columns):
        # 初始宽度：先算表头的长度
        max_width = get_visual_length(col_name)

        # 扫描该列的前 100 行数据 (避免数据量太大卡顿)，找到最长的一个
        # 注意：pandas 读取的空值可能是 None 或 NaN，要转字符串
        column_data = df[col_name].fillna('').astype(str).head(100)

        for val in column_data:
            this_width = get_visual_length(val)
            if this_width > max_width:
                max_width = this_width

        # 加上一点左右余量
        final_width = max_width + 2

        # 设限：最小 12，最大 50 (防止某一行写小作文把列撑爆)
        if final_width < 12: final_width = 12
        if final_width > 50: final_width = 50

        col_letter = get_column_letter(i + 1)
        worksheet.column_dimensions[col_letter].width = final_width

    # --- 3. 设置数据区域样式 (从第二行开始) ---
    # 获取最大行和最大列
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    # 遍历所有数据单元格加边框
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = content_font
            cell.border = thin_border
            cell.alignment = center_align  # 如果喜欢左对齐，改成 left_align

#下载“合同基础信息模板”的接口
# app.py

@app.route('/download_contract_template')
@login_required
def download_contract_template():
    # 创建示例数据
    data = {
        '序号': ['1', '2'],
        '合同编号': ['HT2023001', 'HT2023002'],
        '合同名称': ['大数据平台建设合同', '云服务器租赁协议'],
        '甲方单位': ['某某科技公司', '某某集团'],
        '乙方单位': ['xx数据服务商', 'xx云厂商'],
        '合同金额': [150000.00, 50000.00],
        '签约日期': ['2023-01-10', '2023-03-15'],
        '履约日期': ['2023-12-31', '2024-03-14'],
        '扫描件链接': ['', 'http://example.com/file.pdf']
    }
    df = pd.DataFrame(data)

    # 写入内存
    output = io.BytesIO()
    # 调用美化函数
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='合同导入模板')
        beautify_excel(writer, df, sheet_name='合同导入模板')

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='合同信息导入模板.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
# 下载“合同余额模板”的接口

@app.route('/download_balance_template')
@login_required
def download_balance_template():
    # 创建示例数据
    data = {
        '序号': ['1', '2'],
        '合同编号': ['HT20230101', 'HT20230102'],
        '支付金额': [50000.00, 12000.50],
        '支付时间': ['2023-01-15', '2023-02-20'],
        '备注': ['首付款', '进度款']
    }
    df = pd.DataFrame(data)

    # 写入内存
    output = io.BytesIO()
    # 调用美化函数 (确保您之前已经添加了 beautify_excel 函数)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='余额导入模板')
        beautify_excel(writer, df, sheet_name='余额导入模板')

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='合同余额导入模板.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

#---增加下载主题模板接口
@app.route('/download_topic_template')
@login_required
def download_topic_template():
    # 创建示例数据
    data = {
        '序号':['01','02'],
        '一级主题': ['社会建设', '生态文明'],
        '二级主题': ['教育文化', '环境保护'],
        '主题ID': ['A001', 'B002'],
        '主题名称': ['高校科研数据', '水质监测数据'],
        '采集方式': ['程序采集', '人工采集,程序采集'],
        '采集频率': ['周', '月'],
        '负责人': ['张三', '李四']
    }
    df = pd.DataFrame(data)

    # 写入内存
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='主题导入模板')
        beautify_excel(writer, df, sheet_name='主题导入模板')
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='主题导入模板.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# app.py
#---增加下载服务模板接口
@app.route('/download_task_template')
@login_required
def download_task_template():
    # 创建服务内容导入的示例数据
    data = {
        '序号':['1','2'],
        '合同名称': ['XX省数据采购项目', '智慧城市建设合同'],
        '服务内容': ['提供全省高校科研数据采集服务', '提供实时水质监测API接口'],
        '主题名称': ['科研数据', '水质监测']
    }
    df = pd.DataFrame(data)

    # 写入内存
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='服务内容导入模板')
        beautify_excel(writer, df, sheet_name='服务内容导入模板')
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='服务内容导入模板.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
@app.route('/api/get_level2_topics')
@login_required
def get_level2_topics():
    level1_id = request.args.get('level1_id', '').strip()
    if not level1_id:
        # 如果没选一级，返回所有二级主题
        level2_list = db.session.query(Topic.level2_id).distinct().all()
    else:
        # 如果选了一级，只返回属于该一级的二级主题
        level2_list = db.session.query(Topic.level2_id).filter(Topic.level1_id == level1_id).distinct().all()

    return jsonify([t[0] for t in level2_list if t[0]])


@app.route('/api/get_topic_relations')
@login_required
def get_topic_relations():
    # 获取当前选中的值
    level1 = request.args.get('level1', '').strip()
    level2 = request.args.get('level2', '').strip()
    name = request.args.get('name', '').strip()

    # 从数据库查询现有的关联组合
    query = db.session.query(Topic.level1_id, Topic.level2_id, Topic.name)

    if level1:
        query = query.filter(Topic.level1_id == level1)
    if level2:
        query = query.filter(Topic.level2_id == level2)
    if name:
        query = query.filter(Topic.name == name)

    results = query.distinct().all()

    # 返回所有合法的关联项列表
    return jsonify({
        'level1_list': list(set([r[0] for r in results if r[0]])),
        'level2_list': list(set([r[1] for r in results if r[1]])),
        "name_list": list(set([r[2] for r in results if r[2]]))
    })


# === 任务管理 ===
# app.py 中的 tasks 函数
@app.route('/tasks')
@login_required
def tasks():
    page = request.args.get('page', 1, type=int)

    # 1. 获取搜索参数
    s_contract = request.args.get('contract_name', '').strip()
    s_service = request.args.get('service_content', '').strip()
    s_theme = request.args.get('theme_name', '').strip()

    # 新增的关联搜索参数
    s_code = request.args.get('contract_code', '').strip()
    s_party_a = request.args.get('party_a', '').strip()
    s_party_b = request.args.get('party_b', '').strip()

    # 2. 构建查询 (保持 outerjoin Contract 以支持合同搜索)
    query = Task.query.outerjoin(Contract).options(joinedload(Task.contract))

    # ... (中间的 filter 过滤逻辑保持不变) ...
    if s_contract: query = query.filter(Task.contract_name.contains(s_contract))
    if s_service: query = query.filter(Task.service_content.contains(s_service))
    if s_theme: query = query.filter(Task.theme_name.contains(s_theme))
    if s_code: query = query.filter(Contract.contract_code.contains(s_code))
    if s_party_a: query = query.filter(Contract.party_a.contains(s_party_a))
    if s_party_b: query = query.filter(Contract.party_b.contains(s_party_b))

    pagination = query.order_by(Task.id.desc()).paginate(page=page, per_page=50)

    # ==========================================
    # 🟢【新增】构建 "主题名称 -> 主题ID" 的字典
    # ==========================================
    # 查出所有主题的 (名字, ID)
    all_topics = Topic.query.with_entities(Topic.name, Topic.theme_id).all()
    # 生成字典，例如：{'文物采集': 'A-01', '数据清洗': 'B-02'}
    # 加上 if t.name 判断防止报错
    topic_map = {t.name: t.theme_id for t in all_topics if t.name}

    return render_template('match.html',
                           title="服务内容",
                           pagination=pagination,
                           topic_map=topic_map)  # 👈 记得把字典传给前端
@app.route('/task/edit/<int:id>', methods=['GET', 'POST'])
@app.route('/task/add', methods=['GET', 'POST'], defaults={'id': None})
@login_required
def edit_task(id):
    task = Task.query.get_or_404(id) if id else None
    if request.method == 'POST':
        if not task:
            task = Task()
            db.session.add(task)

        # 1. 获取表单数据
        c_name = request.form.get('contract_name', '').strip()
        task.contract_name = c_name
        task.service_content = request.form.get('service_content')
        task.theme_name = request.form.get('theme_name')

        # 2. 🟢【新增】手动编辑时，也要自动查找并关联 ID
        if c_name:
            contract = Contract.query.filter_by(name=c_name).first()
            if contract:
                task.contract_id = contract.id
            else:
                # 如果名字改错了，或者改成了不存在的合同，要把 ID 清空，否则会关联到错误的合同
                task.contract_id = None
        else:
            task.contract_id = None

        db.session.commit()
        flash('服务内容已保存', 'success')
        return redirect(url_for('tasks'))
    # 获取所有合同对象，包含编号、甲方、乙方信息
    all_contracts = Contract.query.order_by(Contract.id.desc()).all()
    all_topics = Topic.query.order_by(Topic.id.desc()).all()  # ✨ 新增：获取所有主题

    # 转换成简单列表 ['合同A', '合同B', ...]
    contract_list = [c.name for c in all_contracts]

    return render_template('match_form.html',
                           task=task,
                           contract_list=contract_list,
                           all_contracts=all_contracts,
                           all_topics=all_topics)


@app.route('/task/delete/<int:id>')
@login_required
def delete_task(id):
    db.session.delete(Task.query.get_or_404(id))
    db.session.commit()
    return redirect(url_for('tasks'))


@app.route('/tasks/add_manual', methods=['POST'])
@login_required
def add_task_manual():
    try:
        # ... 获取表单数据 ...
        name_from_form = request.form.get('contract_name')
        new_task = Task(
            contract_name=request.form.get('contract_name'),
            service_content=request.form.get('service_content'),
            theme_name=request.form.get('theme_name')
        )
        if name_from_form:
            contract = Contract.query.filter_by(name=name_from_form).first()
            if contract:
                new_task.contract_id = contract.id
        db.session.add(new_task)
        db.session.commit()
        flash('添加成功', 'success')
    except Exception as e:
        #flash(f'添加失败: {e}', 'danger')
        flash('添加失败: {}'.format(e), 'danger')
    return redirect(url_for('tasks'))


@app.route('/import_task_excel', methods=['POST'])
@login_required
def import_task_excel():
    file = request.files.get('file')
    if not file: return "无文件", 400
    try:
        # 读取 Excel
        df = pd.read_excel(file).replace({pd.NA: None, float('nan'): None})

        success_count = 0

        for _, row in df.iterrows():
            # 1. 获取并清洗 Excel 数据
            c_name = str(row.get('合同名称') or '').strip()
            s_content = str(row.get('服务内容') or '').strip()
            t_name = str(row.get('主题名称') or '').strip()

            if not s_content:  # 如果没有服务内容，跳过
                continue

            # 2. 创建唯一的任务对象
            task = Task(
                contract_name=c_name,
                service_content=s_content,
                theme_name=t_name
            )

            # 3. 自动关联逻辑：根据名字找 ID
            if c_name:
                # 去合同表里查，有没有叫这个名字的？
                linked_contract = Contract.query.filter_by(name=c_name).first()
                if linked_contract:
                    task.contract_id = linked_contract.id  # 找到了！绑上去！
                else:
                    task.contract_id = None  # 没找到，保持为空
                    # print(f"警告：未找到名称为 {c_name} 的合同")

            # 4. 只添加这一个对象
            db.session.add(task)
            success_count += 1

        db.session.commit()
        # flash(f'导入完成，成功添加 {success_count} 条任务', 'success')
        flash('导入完成，成功添加 {} 条任务'.format(success_count), 'success')
        return redirect(url_for('tasks'))

    except Exception as e:
        db.session.rollback()
        # return f"错误: {e}", 500
        return "错误: {}".format(e), 500
# === 批量删除 ===
# --- 修正后的批量删除合同 (app.py) ---

@app.route('/contracts/batch_delete', methods=['POST'])
@login_required
def batch_delete_contracts():
    # 改為接收 JSON
    data = request.get_json()
    ids = data.get('ids', []) if data else []

    if not ids:
        return jsonify({'status': 'error', 'message': '未勾選合同'}), 400

    try:
        contracts_to_delete = Contract.query.filter(Contract.id.in_(ids)).all()
        count = len(contracts_to_delete)
        for c in contracts_to_delete:
            db.session.delete(c)
        db.session.commit()
        return jsonify({'status': 'success', 'message': f'成功刪除 {count} 份合同'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/topics/batch_delete', methods=['POST'])
@login_required
def batch_delete_topics():
    ids = request.form.getlist('selected_ids')
    if ids:
        Topic.query.filter(Topic.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
       # flash(f'已删除 {len(ids)} 条主题', 'success')
        flash('已删除 {} 条主题'.format(len(ids)), 'success')
    return redirect(url_for('topics'))


@app.route('/tasks/batch_delete', methods=['POST'])
@login_required
def batch_delete_tasks():
    ids = request.form.getlist('selected_ids')
    if ids:
        Task.query.filter(Task.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        #flash(f'已删除 {len(ids)} 条任务', 'success')
        flash('已删除 {} 条任务'.format(len(ids)), 'success')
    return redirect(url_for('tasks'))


# --- 新增：服务内容与主题对应视图 ---
# app.py
# --- app.py 修改片段 ---

@app.route('/import_service_mapping', methods=['POST'])
@login_required
def import_service_mapping():
    file = request.files.get('file')
    if not file:
        flash("未选择文件", "danger")
        return redirect(url_for('service_mapping'))

    try:
        # 1. 读取并预处理：将所有 NaN 替换为 None，并将所有列转为字符串
        df = pd.read_excel(file).fillna('').astype(str)

        # 检查表头是否存在 (请确保 Excel 表头与此处文字一致)
        required_cols = ['合同名称', '服务内容', '主题名称']
        if not all(col in df.columns for col in required_cols):
           # flash(f"Excel 格式错误！必须包含列：{', '.join(required_cols)}", "danger")
            flash("Excel 格式错误！必须包含列：{}".format(', '.join(required_cols)), "danger")
            return redirect(url_for('service_mapping'))

        import_count = 0
        skip_count = 0

        for _, row in df.iterrows():
            # 2. 强力清洗：去除首尾空格、去除中间换行符
            c_name = row.get('合同名称', '').strip().replace('\n', '').replace('\r', '')
            s_content = row.get('服务内容', '').strip().replace('\n', '').replace('\r', '')
            t_name = row.get('主题名称', '').strip().replace('\n', '').replace('\r', '')

            if not c_name or not s_content or not t_name:
                continue

            # --- 核心查重邏輯 ---
            # 檢查數據庫中是否已存在完全相同的記錄
            exists = ServiceMapping.query.filter_by(
                contract_name=c_name,
                service_content=s_content,
                theme_name=t_name
            ).first()

            if exists:
                skip_count += 1
                continue  # 如果存在，跳過此行
            # ------------------
            # 直接新增对应关系，支持一个服务对应多个主题
            new_map = ServiceMapping(
                contract_name=c_name,
                service_content=s_content,
                theme_name=t_name
            )
            db.session.add(new_map)
            import_count += 1

        db.session.commit()
        # 反饋導入結果，包含成功數和跳過數
        #msg = f"導入完成！成功新增 {import_count} 條。"
        msg = "導入完成！成功新增 {} 條。".format(import_count)
        if skip_count > 0:
            #msg += f"（檢測到 {skip_count} 條重複數據已自動跳過）"
            msg += "（檢測到 {} 條重複數據已自動跳過）".format(skip_count)

        flash(msg, "success" if import_count > 0 else "info")
    except Exception as e:
        db.session.rollback()
        #flash(f"导入失败：{str(e)}", "danger")
        flash("导入失败：{}".format(str(e)), "danger")

    return redirect(url_for('service_mapping'))


# app.py - 完全替换 service_mapping 函数
@app.route('/service_mapping')
@login_required
def service_mapping():
    # 1. 获取筛选参数
    s_contract = request.args.get('contract', '').strip()
    expand_c = request.args.get('expand_c', '').strip()
    expand_s = request.args.get('expand_s', '').strip()

    # 2. 【核心变化】直接查询 Task 表 (服务内容表)
    # 使用 joinedload 预加载合同信息，防止 N+1 查询
    query = Task.query.options(joinedload(Task.contract))

    if s_contract:
        query = query.filter(Task.contract_name.contains(s_contract))

    # 按合同名称排序，保证显示顺序
    all_tasks = query.order_by(Task.contract_name, Task.id.desc()).all()

    # 3. 【数据重构】将扁平的 Task 数据转换为三级嵌套结构
    # 结构目标: grouped_data[合同名][服务内容] = [Task对象列表]
    grouped_data = {}
    contract_theme_counts = {}  # 统计每个合同下的主题数

    for task in all_tasks:
        # 获取名称 (优先用关联对象的名称，如果为空则用 task 表存的快照)
        c_name = task.contract.name if task.contract else (task.contract_name or "未关联/未命名合同")
        s_content = task.service_content or "未填写服务内容"

        # 初始化字典层级
        if c_name not in grouped_data:
            grouped_data[c_name] = {}
        if s_content not in grouped_data[c_name]:
            grouped_data[c_name][s_content] = []

        # 将 task 对象放入对应的格子
        grouped_data[c_name][s_content].append(task)

        # 统计计数
        contract_theme_counts[c_name] = contract_theme_counts.get(c_name, 0) + 1

    # 4. 构建合同详情字典 (用于在折叠栏显示编号、甲乙方)
    # 直接查询所有合同，建立 "名称 -> 对象" 的索引
    all_contracts_db = Contract.query.all()
    contract_map = {c.name: c for c in all_contracts_db}

    # 5. 获取其他辅助数据 (用于下拉框等)
    all_topics = Topic.query.all()
    all_parents = [t[0] for t in db.session.query(Topic.level1_id).distinct().all() if t[0]]
    all_level2_list = [t[0] for t in db.session.query(Topic.level2_id).distinct().all() if t[0]]
    all_topic_names = [t[0] for t in db.session.query(Topic.name).distinct().all() if t[0]]

    raw_relations = db.session.query(Topic.level1_id, Topic.level2_id, Topic.name).distinct().all()
    topic_relations = [{"p": r[0] or "", "l2": r[1] or "", "n": r[2] or ""} for r in raw_relations]

    return render_template('service.html',
                           grouped_data=grouped_data,  # 现在这里面装的是 Task 对象
                           theme_counts=contract_theme_counts,
                           contract_map=contract_map,  # 合同详情字典
                           all_topics=all_topics,
                           all_parents=all_parents,
                           all_level2_list=all_level2_list,
                           all_topic_names=all_topic_names,
                           topic_relations=topic_relations,
                           expand_c=expand_c,
                           expand_s=expand_s)

class ServiceMapping(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    contract_name = db.Column(db.String(255))  # 一级：合同名称
    service_content = db.Column(db.String(255))  # 二级：服务内容
    theme_name = db.Column(db.String(255))  # 三级：主题名称


@app.route('/unified_delete')
@login_required
def unified_delete():
    target_type = request.args.get('type')  # 'contract' 或 'mapping'
    target_id = request.args.get('id')  # 对应的主键 ID 或 合同名称

    try:
        if target_type == 'contract':
            # 批量删除该合同下的所有记录
            ServiceMapping.query.filter_by(contract_name=target_id).delete()
            #flash(f"合同「{target_id}」及关联内容已清理", "success")
            flash("合同「{}」及关联内容已清理".format(target_id), "success")

        elif target_type == 'mapping':
            # 删除单条对应关系
            item = ServiceMapping.query.get_or_404(target_id)
            db.session.delete(item)
            flash("单条对应关系已删除", "info")

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        #flash(f"操作失败: {str(e)}", "danger")
        flash("操作失败: {}".format(str(e)), "danger")

    return redirect(url_for('service_mapping'))


@app.route('/delete_mapping/<int:id>')
@login_required
def delete_mapping(id):
    # 【核心】现在删除的是 Task 对象
    task = Task.query.get_or_404(id)
    c_name = task.contract_name
    s_content = task.service_content

    try:
        db.session.delete(task)
        db.session.commit()
        flash("关联已删除", "success")
    except Exception as e:
        db.session.rollback()
        flash("删除失败: {}".format(str(e)), "danger")

    return redirect(url_for('service_mapping', expand_c=c_name, expand_s=s_content))
@app.route('/remove_contract_group', methods=['POST'])
@login_required
def remove_contract_group():
    target_contract = request.form.get('contract_name')
    if not target_contract:
        flash("参数错误", "danger")
        return redirect(url_for('service_mapping'))

    try:
        # 【核心】从 Task 表中删除该合同的所有服务记录
        num_deleted = Task.query.filter(Task.contract_name == target_contract).delete()
        db.session.commit()
        flash("已清理合同「{}」，共删除 {} 条服务记录".format(target_contract, num_deleted), "success")
    except Exception as e:
        db.session.rollback()
        flash("系统错误: {}".format(str(e)), "danger")

    return redirect(url_for('service_mapping'))
# 修改完模型后，记得在命令行运行 db.create_all() 或重启时自动创建
@app.route('/add_service_mapping', methods=['POST'])
@login_required
def add_service_mapping():
    # 1. 获取表单数据
    c_name = request.form.get('contract_name', '').strip()
    s_content = request.form.get('service_content', '').strip()
    theme_names = request.form.getlist('theme_names')  # 多选主题

    if not c_name or not s_content or not theme_names:
        flash("所有字段均为必填项", "warning")
        return redirect(url_for('service_mapping'))

    # 2. 查找关联的合同ID (自动关联逻辑)
    contract_id = None
    linked_contract = Contract.query.filter_by(name=c_name).first()
    if linked_contract:
        contract_id = linked_contract.id

    success_count = 0
    duplicate_count = 0

    try:
        for t_name in theme_names:
            if not t_name.strip(): continue
            t_name = t_name.strip()

            # 3. 查重 (检查 Task 表)
            exists = Task.query.filter_by(
                contract_name=c_name,
                service_content=s_content,
                theme_name=t_name
            ).first()

            if exists:
                duplicate_count += 1
                continue

            # 4. 【核心】创建 Task 对象
            new_task = Task(
                contract_name=c_name,
                service_content=s_content,
                theme_name=t_name,
                contract_id=contract_id  # 自动填入ID
            )
            db.session.add(new_task)
            success_count += 1

        db.session.commit()

        if success_count > 0:
            flash("成功新增 {} 个服务-主题关联".format(success_count), "success")
        elif duplicate_count > 0:
            flash("所选关联已存在，未重复添加", "info")

    except Exception as e:
        db.session.rollback()
        flash("新增失败：{}".format(str(e)), "danger")

    # 保持页面展开状态
    return redirect(url_for('service_mapping', expand_c=c_name, expand_s=s_content))

# 在app.py中添加API端点
@app.route('/api/get_topics', methods=['GET'])
@login_required
def get_topics():
    """获取所有主题数据，用于AJAX请求"""
    topics = Topic.query.order_by(Topic.name).all()
    topics_data = []
    for topic in topics:
        topics_data.append({
            'id': topic.id,
            'name': topic.name,
            'theme_id': topic.theme_id,
            'level1': topic.level1_id,
            'level2': topic.level2_id,
            'owner': topic.owner
        })
    return jsonify(topics_data)


# app.py
#---------------任务管理--------------------

# app.py

# --- 1. 修改模型：增加 data_count 字段 ---
class TaskInstance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.Integer, db.ForeignKey('task.id'))
    #新增这行关联代码通过 task_instance.task 访问原始任务信息了
    task = db.relationship('Task')
    contract_name = db.Column(db.String(200))
    theme_name = db.Column(db.String(200))
    owner = db.Column(db.String(100))

    belong_month = db.Column(db.String(20))  # 所属年月
    deadline = db.Column(db.String(20))  # 截止时间

    #新增字段
    status = db.Column(db.String(20), default='pending')  # pending(进行中) / completed(已完成)
    data_count = db.Column(db.Integer, default=0)  # 任务数据量

    #新增字段：实际完成时间
    finished_at = db.Column(db.String(20))
    created_at = db.Column(db.DateTime, default=datetime.now)


# --- 2. 修改路由：加载分页所需的所有数据 ---
# app.py -> task_management 函数 (完全替换)
@app.route('/task_management')
@login_required
def task_management():
    # ==================================================
    # 0. 预处理：获取所有负责人名单 (提至最前，供 Tab 2 和 Tab 3 共用)
    # ==================================================
    # 获取配置中的负责人
    config_owners = set(get_all_owners())

    # 获取任务实例表中实际存在的负责人 (防止历史数据里的负责人不在配置表中)
    active_owners_query = db.session.query(TaskInstance.owner).distinct().all()
    active_owners = set([r[0] for r in active_owners_query if r[0]])

    # 合并并排序
    all_owners = sorted(list(config_owners | active_owners))
    # 确保 '(未指定负责人)' 在列表最后
    if '(未指定负责人)' in all_owners:
        all_owners.remove('(未指定负责人)')
        all_owners.append('(未指定负责人)')

    # ==================================================
    # Tab 1: 任务生成 (搜索逻辑 - 保持不变)
    # ==================================================
    q = request.args.get('q', '').strip()
    search_results = []
    if q:
        query = db.session.query(
            Task.id, Task.contract_name, Task.service_content, Task.theme_name,
            Contract.contract_code, Topic.theme_id, Topic.frequency, Topic.owner
        ).outerjoin(Contract, Task.contract_id == Contract.id) \
            .outerjoin(Topic, Task.theme_name == Topic.name)

        search_filter = or_(
            Task.theme_name.contains(q), Topic.theme_id.contains(q),
            Task.contract_name.contains(q), Contract.contract_code.contains(q)
        )
        raw_data = query.filter(search_filter).all()
        for item in raw_data:
            search_results.append({
                'id': item.id, 'contract_name': item.contract_name,
                'service_content': item.service_content, 'theme_name': item.theme_name,
                'theme_id': item.theme_id, 'frequency': item.frequency, 'owner': item.owner
            })

    # ==================================================
    # Tab 2: 任务查看 (✨ 修复：支持4字段组合筛选)
    # ==================================================
    # 1. 获取参数
    view_contract = request.args.get('view_contract', '').strip()
    view_theme = request.args.get('view_theme', '').strip()
    view_owner = request.args.get('view_owner', '').strip()
    view_status = request.args.get('view_status', 'all')

    instance_query = TaskInstance.query

    # 2. 逐个应用筛选 (支持模糊搜索)
    if view_contract:
        instance_query = instance_query.filter(TaskInstance.contract_name.contains(view_contract))
    if view_theme:
        instance_query = instance_query.filter(TaskInstance.theme_name.contains(view_theme))
    if view_owner and view_owner != 'all':
        instance_query = instance_query.filter(TaskInstance.owner == view_owner)
    if view_status != 'all':
        instance_query = instance_query.filter_by(status=view_status)

    # 3. 执行查询
    all_instances = instance_query.order_by(TaskInstance.deadline.asc()).all()
    ongoing_count = TaskInstance.query.filter_by(status='pending').count()

    # ==================================================
    # Tab 3: 任务统计 (逻辑保持不变，但共用顶部的 all_owners)
    # ==================================================
    stat_month = request.args.get('month', datetime.now().strftime('%Y-%m'))

    try:
        y, m = map(int, stat_month.split('-'))
        _, num_days = calendar.monthrange(y, m)
        date_list = ["{}-{:02d}-{:02d}".format(y, m, d) for d in range(1, num_days + 1)]
    except:
        date_list = []

        # 查询统计数据
    raw_stats = db.session.query(
        TaskInstance.finished_at,
        TaskInstance.owner,
        func.sum(TaskInstance.data_count)
    ).filter(
        TaskInstance.status == 'completed',
        TaskInstance.finished_at.startswith(stat_month)
    ).group_by(TaskInstance.finished_at, TaskInstance.owner).all()

    # 动态日期去重
    active_dates = set()
    for r in raw_stats:
        if r[0]: active_dates.add(r[0])
    date_list = sorted(list(active_dates))

    # 填入矩阵
    stats_matrix = {d: {o: 0 for o in all_owners} for d in date_list}
    row_totals = {d: 0 for d in date_list}
    col_totals = {o: 0 for o in all_owners}
    grand_total = 0

    for date_str, owner, count in raw_stats:
        if not count: count = 0
        if date_str in stats_matrix:
            real_owner = owner if owner and owner in col_totals else '(未指定负责人)'
            if real_owner in stats_matrix[date_str]:
                stats_matrix[date_str][real_owner] += count
                row_totals[date_str] += count
                col_totals[real_owner] += count
                grand_total += count

    return render_template('task_management.html',
                           results=search_results,
                           all_instances=all_instances,
                           ongoing_count=ongoing_count,
                           # ✨ 关键：传递筛选参数回显
                           view_contract=view_contract,
                           view_theme=view_theme,
                           view_owner=view_owner,
                           view_status=view_status,
                           # 统计参数
                           stat_month=stat_month,
                           all_owners=all_owners,
                           date_list=date_list,
                           stats_matrix=stats_matrix,
                           row_totals=row_totals,
                           col_totals=col_totals,
                           grand_total=grand_total)

@app.route('/api/update_task_instance', methods=['POST'])
@login_required
def update_task_instance():
    data = request.json
    instance = TaskInstance.query.get(data.get('id'))
    if not instance:
        return jsonify({'status': 'error', 'message': '任务不存在'})

    try:
        # 更新字段
        if 'deadline' in data: instance.deadline = data['deadline']
        if 'status' in data: instance.status = data['status']
        if 'data_count' in data: instance.data_count = int(data['data_count'])
        # ✨ 新增：更新完成时间
        if 'finished_at' in data: instance.finished_at = data['finished_at']
        # ✨ 新增：允许修改负责人
        if 'owner' in data: instance.owner = data['owner']
        db.session.commit()
        return jsonify({'status': 'success', 'message': '任务更新成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)})

# ... (在 update_task_instance 函数下方添加) ...
@app.route('/api/delete_task_instance/<int:id>', methods=['POST'])
@login_required
def delete_task_instance(id):
    try:
        # 获取要删除的任务实例
        instance = TaskInstance.query.get_or_404(id)

        # 执行删除
        db.session.delete(instance)
        db.session.commit()

        return jsonify({'status': 'success', 'message': '任务已成功删除'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': '删除失败: {}'.format(str(e))})

# --- 3. 新增：生成任务的保存接口 ---
@app.route('/api/generate_task', methods=['POST'])
@login_required
def generate_task_api():
    data = request.json
    task_id = data.get('task_id')

    try:
        # 查重：防止同一个月重复生成同一个任务
        exists = TaskInstance.query.filter_by(
            task_id=task_id,
            belong_month=data.get('belong_month')
        ).first()

        if exists:
            return jsonify({'status': 'warning', 'message': '该任务本月已生成过，无需重复生成！'})

        new_instance = TaskInstance(
            task_id=task_id,
            contract_name=data.get('contract_name'),
            theme_name=data.get('theme_name'),
            owner=data.get('owner'),
            belong_month=data.get('belong_month'),
            deadline=data.get('deadline')
        )
        db.session.add(new_instance)
        db.session.commit()
        return jsonify({'status': 'success', 'message': '任务已成功下发！'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)})


# app.py

# ... (在 generate_task_api 函数下方添加) ...

@app.route('/api/batch_generate_tasks', methods=['POST'])
@login_required
def batch_generate_tasks():
    data = request.json
    tasks_data = data.get('tasks', [])

    if not tasks_data:
        return jsonify({'status': 'error', 'message': '未接收到任务数据'})

    success_count = 0
    skipped_count = 0

    try:
        for item in tasks_data:
            # 1. 查重：同一任务ID在同月是否已存在
            exists = TaskInstance.query.filter_by(
                task_id=item.get('task_id'),
                belong_month=item.get('belong_month')
            ).first()

            if exists:
                skipped_count += 1
                continue

            # 2. 创建新任务实例
            new_instance = TaskInstance(
                task_id=item.get('task_id'),
                contract_name=item.get('contract_name'),
                theme_name=item.get('theme_name'),
                owner=item.get('owner') or '(未指定负责人)',
                belong_month=item.get('belong_month'),
                deadline=item.get('deadline'),
                status='pending',
                data_count=0
            )
            db.session.add(new_instance)
            success_count += 1

        db.session.commit()

        msg = f"操作完成：成功生成 {success_count} 条"
        if skipped_count > 0:
            msg += f"，跳过 {skipped_count} 条重复记录"

        return jsonify({'status': 'success', 'message': msg})

    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)})
# app.py -> 新增工具路由

@app.route('/admin/fix_empty_owners')
@login_required
def fix_empty_owners():
    """
    数据清洗工具：将所有没有负责人的主题，统一修正为 '(未指定负责人)'
    """
    try:
        # 查找所有：NULL, 空字符串, 或 错误的 'None' 字符串
        affected_rows = Topic.query.filter(
            or_(
                Topic.owner == None,
                Topic.owner == '',
                Topic.owner == 'None'
            )
        ).update({Topic.owner: "(未指定负责人)"}, synchronize_session=False)

        db.session.commit()
        flash(f'✅ 数据清洗完成！已将 {affected_rows} 条无主数据的负责人修正为“(未指定负责人)”。', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ 修复失败: {str(e)}', 'danger')

    return redirect(url_for('topics'))


# app.py

# ... (在 delete_task_instance 函数下方添加) ...

@app.route('/api/batch_delete_task_instances', methods=['POST'])
@login_required
def batch_delete_task_instances():
    data = request.json
    ids = data.get('ids', [])

    if not ids:
        return jsonify({'status': 'error', 'message': '未选择任何任务'})

    try:
        # 批量删除 (使用 in_ 查询)
        # synchronize_session=False 可以提高删除效率
        num_deleted = TaskInstance.query.filter(TaskInstance.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()

        return jsonify({'status': 'success', 'message': f'成功删除 {num_deleted} 条任务'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)})

#导出合同Excel
@app.route('/api/export_contracts', methods=['POST'])
@login_required
def export_contracts():
    try:
        data = request.json
        ids = data.get('ids', [])
        contracts = Contract.query.filter(Contract.id.in_(ids)).all()

        # 1. 构造带序号的数据
        export_data = []
        for i, c in enumerate(contracts, 1):
            export_data.append({
                '序号': i,
                '合同编号': getattr(c, 'contract_code', ''),
                '合同名称': getattr(c, 'name', ''),  # 请确保字段名对应数据库
                '甲方单位': getattr(c, 'party_a', '')
            })

        df = pd.DataFrame(export_data)
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            ws = writer.sheets['Sheet1']

            # 样式美化
            header_fill = PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                            bottom=Side(style='thin'))
            alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 列宽设定
            ws.column_dimensions['A'].width = 8  # 序号
            ws.column_dimensions['B'].width = 20  # 编号
            ws.column_dimensions['C'].width = 50  # 名称列较长
            ws.column_dimensions['D'].width = 50
            # 2. 遍历数据行：设置行高与自动换行
            for row in range(1, len(export_data) + 2):
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = alignment
                    if row == 1:
                        cell.fill = header_fill
                        cell.font = Font(bold=True)

                if row > 1:
                    content = str(export_data[row - 2]['合同名称'])
                    # 计算显示长度：中文2位，英数1位
                    actual_len = sum(2 if ord(char) > 127 else 1 for char in content)
                    # 根据列宽50（实际内容区约46）计算行数
                    lines = (actual_len // 46) + 1
                    ws.row_dimensions[row].height = lines * 20  # 动态行高

        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f"合同导出_{datetime.now().strftime('%m%d')}.xlsx")
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
# --- 7. 程序启动 ---
# app.py 末尾
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        # 检查是否已有管理员
        admin = User.query.filter_by(username='admin').first()
        if not admin:
            admin = User(username='admin')
            # 使用兼容模式设置密码
            admin.password_hash = generate_password_hash('123456', method='pbkdf2:sha256')
            db.session.add(admin)
            db.session.commit()
            print("管理员账号已创建: admin / 123456")

    app.run(host='0.0.0.0', port=1027)
